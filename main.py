import os
import re
import shutil
import sys
from pathlib import Path

import numpy as np
import pandas as pd
import yaml
from Levenshtein import ratio
from PySide6.QtCore import Slot
from PySide6.QtWidgets import (
    QApplication,
    QFileDialog,
    QHBoxLayout,
    QLabel,
    QMainWindow,
    QPushButton,
    QVBoxLayout,
    QWidget,
)
from openpyxl.reader.excel import load_workbook


def get_config() -> dict:
    with open("config.yaml", "r") as file:
        config = yaml.safe_load(file)
    return config


config = get_config()

CURRENT_PATH = Path(os.getcwd())
INFO_PATH = Path(config.get("info_path") or CURRENT_PATH / "zp_file" / "Расчет ЗП.xlsx")
FROM_FILES_PATH = Path(config.get("files_path") or CURRENT_PATH / "files")

TO_FILES_PATH = Path(config.get("files_new_path") or CURRENT_PATH / "files_new")
PASSWD = str(config.get("password"))
SIMILARITY_RATIO = 0.8

SPECIALIZATION_MAP = {
    "МАНИКЮР": "МАНИКЮР-ПЕДИКЮР",
    "ПЕДИКЮР": "МАНИКЮР-ПЕДИКЮР",
    "ВИЗАЖ": "РЕСНИЦЫ, ВИЗАЖ",
    "РЕСНИЦЫ": "РЕСНИЦЫ, ВИЗАЖ",
    "МАССАЖ": "МАССАЖ лица",
}


def get_zp_df() -> pd.DataFrame:
    def split_empls(empls_value: str) -> list[str]:
        empls_value = [empls_value]
        for key in ["\n", ",", "  "]:
            empls_value = [x for emp in empls_value for x in emp.split(key)]
        return empls_value

    df = pd.read_excel(INFO_PATH, header=1, sheet_name="расчет ЗП")
    df[["Правило", "Сотрудник"]] = df[["Правило", "Сотрудник"]].ffill()
    # new_df = pd.DataFrame(columns=df.columns)
    rows = []

    for _, row in df.iterrows():
        empls = split_empls(row["Сотрудник"])

        for emp in empls:
            if emp:
                emp = emp.strip()
                try:
                    spec = row["Специализация"].split("\n")
                except AttributeError:
                    spec = [row["Специализация"]]
                for s in spec:
                    if not s:
                        continue
                    s = str(s).strip()
                    new_row = row.copy()
                    new_row["Специализация"] = s
                    new_row["Сотрудник"] = emp
                    rows.append(new_row)
    return pd.DataFrame(rows, columns=df.columns)


def get_files_df() -> list[Path]:
    return [
        file for file in FROM_FILES_PATH.iterdir() if file.suffix in [".xlsx", ".xls"]
    ]


def calc_zp(fl: Path, zp_df: pd.DataFrame) -> None:
    def get_proc_to_zp_dict(zp_df: pd.DataFrame) -> dict:
        proc_to_zp_dict = dict(zip(zp_df["Специализация"], zp_df["Процент в ЗП"]))
        return proc_to_zp_dict

    def keep_only_digits(input_string: str) -> [int, None]:
        try:
            return int(re.sub(r"\D", "", str(input_string)))
        except ValueError:
            return

    def get_closest_match(procedure: str, proc_to_zp: dict) -> [str, None]:
        max_ratio = SIMILARITY_RATIO
        max_proc = None
        for proc in proc_to_zp:
            r = ratio(procedure, proc)
            if r > max_ratio:
                max_ratio = r
                max_proc = proc
        return max_proc

    def get_match(first: str, second: str) -> bool:
        if not first or not second:
            return False
        return ratio(first, second) > SIMILARITY_RATIO

    def convert_xls_to_xlsx(file_path: Path) -> Path:
        if file_path.suffix.lower() != ".xls":
            return file_path

        import xlwings as xw

        xlsx_path = file_path.with_suffix(".xlsx")

        app = xw.App(visible=False)
        try:
            wb = app.books.open(str(file_path))
            wb.save(str(xlsx_path))
            wb.close()
            os.remove(file_path)
        finally:
            app.quit()

        return xlsx_path

    export_fl = TO_FILES_PATH / fl.name
    suffix = export_fl.suffix.lower()
    if suffix == ".xls":
        shutil.copy(fl, export_fl)
        export_fl = convert_xls_to_xlsx(export_fl)
    elif suffix == ".xlsx":
        shutil.copy(fl, export_fl)
    else:
        print(f"File {fl} has unsupported format")
        return

    df = pd.read_excel(export_fl)
    df = df.replace(np.nan, None)
    employee = (export_fl.stem.split(" ")[-1]).upper()

    zp_df = zp_df[
        zp_df["Сотрудник"].apply(lambda x: get_match(employee, x.split(" ")[0]))
    ]
    if zp_df.empty:
        print(f"Employee {employee} not found in the ZP file")
        return

    proc_to_zp = get_proc_to_zp_dict(zp_df)

    zp_row = []
    for _, row in df.iterrows():
        procedure = row.iloc[0]
        quantity = row.iloc[1]
        all_price = row.iloc[6]
        if not procedure or not quantity:
            zp_row.append("")
            continue

        zp = ""

        match procedure:
            case "УСЛУГИ СОТРУДНИКАМ":
                zp = all_price
            case "ТОВАРЫ НА ПРОДАЖУ":
                zp = all_price * 0.05
            case "СТРИЖКИ" | "УКЛАДКИ":
                zp = all_price * 0.5
            case "ОКРАШИВАНИЕ ВОЛОС" | "УХОДЫ ДЛЯ ВОЛОС":
                zp = (all_price - all_price * 0.1) * 0.5
            case "РЕСНИЦЫ" | "ВИЗАЖ":
                zp = all_price * 0.5
            case _:
                mp = proc_to_zp.get(procedure)
                if not mp:
                    try_procedure = SPECIALIZATION_MAP.get(procedure)
                    if try_procedure:
                        mp = proc_to_zp.get(try_procedure)
                if not mp:
                    procedure = get_closest_match(procedure, proc_to_zp)
                    mp = proc_to_zp.get(procedure)
                if isinstance(mp, float):
                    if mp > 100:
                        zp = mp * quantity
                    else:
                        zp = mp * all_price
                elif isinstance(mp, str):
                    if mp := keep_only_digits(mp):
                        zp = mp * quantity

        zp_row.append(zp)

    # Load the existing workbook and sheet
    book = load_workbook(export_fl)
    sheet = book.active
    new_column_index = sheet.max_column + 1

    # Add the new column to the existing sheet
    for idx, value in enumerate(zp_row, start=2):  # Assuming header is in the first row
        sheet.cell(row=idx, column=new_column_index, value=value)

    book.save(export_fl)


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Расчет ЗП")
        self.setGeometry(100, 100, 500, 300)

        self.zp_label = QLabel(f"Файл расчета ЗП: {INFO_PATH}")
        self.from_label = QLabel(f"Папка с файлами: {FROM_FILES_PATH}")
        self.to_label = QLabel(f"Папка для результатов: {TO_FILES_PATH}")
        self.zp_label.setWordWrap(True)

        self.select_zp_button = QPushButton("Выбрать файл ЗП")
        self.select_from_button = QPushButton("Выбрать папку с файлами")
        self.select_to_button = QPushButton("Выбрать папку для результатов")
        self.start_button = QPushButton("Старт")
        self.start_button.clicked.connect(self.on_start_clicked)

        self.select_zp_button.clicked.connect(self.on_select_file)
        self.select_from_button.clicked.connect(self.on_select_from_dir)
        self.select_to_button.clicked.connect(self.on_select_to_dir)
        self.start_button.clicked.connect(self.on_start_clicked)

        main_layout = QVBoxLayout()

        pairs = [
            (self.zp_label, self.select_zp_button),
            (self.from_label, self.select_from_button),
            (self.to_label, self.select_to_button),
        ]

        for label, button in pairs:
            h_layout = QHBoxLayout()
            h_layout.addWidget(label, stretch=1)
            h_layout.addWidget(button)
            main_layout.addLayout(h_layout)

        main_layout.addWidget(self.start_button)

        container = QWidget()
        container.setLayout(main_layout)
        self.setCentralWidget(container)

    @Slot()
    def on_select_file(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Выберите файл расчета ЗП", str(CURRENT_PATH), "Excel files (*.xlsx)"
        )
        if file_path:
            # Обновляем конфиг
            config = get_config()
            config["info_path"] = file_path

            # Сохраняем изменения
            with open("config.yaml", "w") as file:
                yaml.dump(config, file, allow_unicode=True)

            # Обновляем глобальную переменную
            global INFO_PATH
            INFO_PATH = Path(file_path)

    @Slot()
    def on_select_from_dir(self):
        dir_path = QFileDialog.getExistingDirectory(
            self, "Выберите папку с файлами", str(CURRENT_PATH)
        )
        if dir_path:
            config = get_config()
            config["files_path"] = dir_path
            with open("config.yaml", "w") as file:
                yaml.dump(config, file, allow_unicode=True)
            global FROM_FILES_PATH
            FROM_FILES_PATH = Path(dir_path)
            self.from_label.setText(f"Папка с файлами: {FROM_FILES_PATH}")

    @Slot()
    def on_select_to_dir(self):
        dir_path = QFileDialog.getExistingDirectory(
            self, "Выберите папку для результатов", str(CURRENT_PATH)
        )
        if dir_path:
            config = get_config()
            config["files_new_path"] = dir_path
            with open("config.yaml", "w") as file:
                yaml.dump(config, file, allow_unicode=True)
            global TO_FILES_PATH
            TO_FILES_PATH = Path(dir_path)
            self.to_label.setText(f"Папка для результатов: {TO_FILES_PATH}")

    @Slot()
    def on_start_clicked(self):
        zp_df = get_zp_df()

        TO_FILES_PATH.mkdir(parents=True, exist_ok=True)

        files = get_files_df()
        for fl in files:
            calc_zp(fl, zp_df)


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())
