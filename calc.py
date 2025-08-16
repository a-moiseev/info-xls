import os
import re
import shutil
from pathlib import Path

import numpy as np
import pandas as pd
from Levenshtein import ratio
from openpyxl.reader.excel import load_workbook

from config import Config

SPECIALIZATION_MAP = {
    "МАНИКЮР": "МАНИКЮР-ПЕДИКЮР",
    "ПЕДИКЮР": "МАНИКЮР-ПЕДИКЮР",
    "ВИЗАЖ": "РЕСНИЦЫ, ВИЗАЖ",
    "РЕСНИЦЫ": "РЕСНИЦЫ, ВИЗАЖ",
    "МАССАЖ": "МАССАЖ лица",
}


class CalcZP:

    def __init__(self, config: Config):
        self.config = config

        self.zp_df = None
        self.files = None

        self.summary_df = None
        self.periods = set()  # Track unique periods

        self.config.to_files_path.mkdir(parents=True, exist_ok=True)

    def parse_date_period(self, file_path: Path) -> str:
        """Extract date period from employee file."""
        try:
            # Read first few rows to find date string
            df = pd.read_excel(file_path, nrows=5)

            # Look for date pattern in first few rows
            for i in range(min(5, len(df))):
                for col in df.columns:
                    cell_value = str(df.iloc[i][col])
                    # Pattern: "за период с 16.07.2025 по 30.07.2025"
                    pattern = r"с\s+(\d{1,2}\.\d{1,2}\.\d{4})\s+по\s+(\d{1,2}\.\d{1,2}\.\d{4})"
                    match = re.search(pattern, cell_value)
                    if match:
                        start_date, end_date = match.groups()
                        # Format as "16.07-30.07.2025" (same year always)
                        start_parts = start_date.split(".")
                        end_parts = end_date.split(".")
                        return f"{start_parts[0]}.{start_parts[1]}-{end_parts[0]}.{end_parts[1]}.{end_parts[2]}"

            # If no date found, return default
            return "Период не найден"

        except Exception as e:
            print(f"Error parsing date from {file_path}: {e}")
            return "Ошибка даты"

    def get_zp_df(self) -> pd.DataFrame:
        def split_empls(empls_value: str) -> list[str]:
            empls_value = [empls_value]
            for key in ["\n", ",", "  "]:
                empls_value = [x for emp in empls_value for x in emp.split(key)]
            return empls_value

        df = pd.read_excel(self.config.info_path, header=1, sheet_name="расчет ЗП")
        df[["Правило", "Сотрудник"]] = df[["Правило", "Сотрудник"]].ffill()
        rows = []

        for _, row in df.iterrows():
            empls = split_empls(row["Сотрудник"])

            for emp in empls:
                if emp:
                    emp = emp.strip()
                    try:
                        spec = row["Специализация"].split("\n")
                    except AttributeError:
                        spec = [row["Специализация"]]
                    for s in spec:
                        if not s:
                            continue
                        s = str(s).strip()
                        new_row = row.copy()
                        new_row["Специализация"] = s
                        new_row["Сотрудник"] = emp
                        rows.append(new_row)
        return pd.DataFrame(rows, columns=df.columns)

    def get_files_df(self) -> list[Path]:
        return [
            file
            for file in self.config.from_files_path.iterdir()
            if file.suffix in [".xlsx", ".xls"]
        ]

    def calc_zp(self, fl: Path, zp_df: pd.DataFrame, progress_callback=None, file_index=0) -> None:
        def get_proc_to_zp_dict(zp_df: pd.DataFrame) -> dict:
            proc_to_zp_dict = dict(zip(zp_df["Специализация"], zp_df["Процент в ЗП"]))
            return proc_to_zp_dict

        def keep_only_digits(input_string: str) -> [int, None]:
            try:
                return int(re.sub(r"\D", "", str(input_string)))
            except ValueError:
                return

        def get_closest_match(procedure: str, proc_to_zp: dict) -> [str, None]:
            max_ratio = self.config.similarity_ratio
            max_proc = None
            for proc in proc_to_zp:
                r = ratio(procedure, proc)
                if r > max_ratio:
                    max_ratio = r
                    max_proc = proc
            return max_proc

        def get_match(first: str, second: str, similarity_ratio: float = 0.8) -> bool:
            if not first or not second:
                return False
            return ratio(first, second) > similarity_ratio

        def convert_xls_to_xlsx(file_path: Path) -> Path:
            if file_path.suffix.lower() != ".xls":
                return file_path

            import xlwings as xw

            xlsx_path = file_path.with_suffix(".xlsx")

            app = xw.App(visible=False)
            try:
                wb = app.books.open(str(file_path))
                wb.save(str(xlsx_path))
                wb.close()
                os.remove(file_path)
            finally:
                app.quit()

            return xlsx_path

        # Этап 1: Копирование файла
        if progress_callback:
            progress_callback(file_index * 4 + 1)
            
        export_fl = self.config.to_files_path / fl.name
        suffix = export_fl.suffix.lower()
        if suffix == ".xls":
            shutil.copy(fl, export_fl)
            
            # Этап 2: Конвертация .xls в .xlsx
            if progress_callback:
                progress_callback(file_index * 4 + 2)
            export_fl = convert_xls_to_xlsx(export_fl)
        elif suffix == ".xlsx":
            shutil.copy(fl, export_fl)
            # Этап 2: Пропускаем конвертацию для .xlsx
            if progress_callback:
                progress_callback(file_index * 4 + 2)
        else:
            print(f"File {fl} has unsupported format")
            return

        # Этап 3: Обработка данных
        if progress_callback:
            progress_callback(file_index * 4 + 3)
            
        df = pd.read_excel(export_fl)
        df = df.replace(np.nan, None)
        employee = (export_fl.stem.split(" ")[0]).upper()

        zp_df = zp_df[
            zp_df["Сотрудник"].apply(
                lambda x: get_match(
                    employee,
                    x.split(" ")[0],
                    self.config.similarity_ratio,
                ),
            )
        ]
        if zp_df.empty:
            print(f"Employee {employee} not found in the ZP file")
            return

        proc_to_zp = get_proc_to_zp_dict(zp_df)

        zp_row = []
        for _, row in df.iterrows():
            procedure = row.iloc[0]
            quantity = row.iloc[1]
            all_price = row.iloc[6]
            if not procedure or not quantity:
                zp_row.append("")
                continue

            zp = ""

            match procedure:
                case "УСЛУГИ СОТРУДНИКАМ":
                    zp = all_price
                case "ТОВАРЫ НА ПРОДАЖУ":
                    zp = all_price * 0.05
                case "СТРИЖКИ" | "УКЛАДКИ":
                    zp = all_price * 0.5
                case "ОКРАШИВАНИЕ ВОЛОС" | "УХОДЫ ДЛЯ ВОЛОС":
                    zp = (all_price - all_price * 0.1) * 0.5
                case "РЕСНИЦЫ" | "ВИЗАЖ":
                    zp = all_price * 0.5
                case _:
                    mp = proc_to_zp.get(procedure)
                    if not mp:
                        try_procedure = SPECIALIZATION_MAP.get(procedure)
                        if try_procedure:
                            mp = proc_to_zp.get(try_procedure)
                    if not mp:
                        procedure = get_closest_match(procedure, proc_to_zp)
                        mp = proc_to_zp.get(procedure)
                    if isinstance(mp, float):
                        if mp > 100:
                            zp = mp * quantity
                        else:
                            zp = mp * all_price
                    elif isinstance(mp, str):
                        if mp := keep_only_digits(mp):
                            zp = mp * quantity

            zp_row.append(zp)

        # Load the existing workbook and sheet
        book = load_workbook(export_fl)
        sheet = book.active
        new_column_index = sheet.max_column + 1

        # Add the new column to the existing sheet
        for idx, value in enumerate(
            zp_row, start=2
        ):  # Assuming header is in the first row
            sheet.cell(row=idx, column=new_column_index, value=value)

        # Удаляем столбцы E, G и H
        cols_to_delete = ["H", "G", "E"]
        for col in cols_to_delete:
            sheet.delete_cols(sheet[col + "1"].column)
        new_column_index = new_column_index - 3

        # Calculate sum of salary values
        total_salary = sum(val for val in zp_row if isinstance(val, (int, float)))

        sum_formula = f"=SUM({sheet.cell(row=2, column=new_column_index).coordinate}:{sheet.cell(row=len(zp_row), column=new_column_index).coordinate})"
        sheet.cell(row=len(zp_row) + 1, column=new_column_index, value=sum_formula)

        # Этап 4: Сохранение файла
        if progress_callback:
            progress_callback(file_index * 4 + 4)
            
        book.save(export_fl)

        # Parse date period and add to summary
        period = self.parse_date_period(fl)
        self.periods.add(period)
        self.add_to_summary(employee, total_salary, period)

    def add_to_summary(self, employee: str, total_salary: float, period: str):
        """Add employee and their total salary to summary DataFrame."""
        employee_name = employee.capitalize()

        # Check if employee already exists
        if employee_name in self.summary_df["Сотрудник"].values:
            # Update existing row
            row_idx = self.summary_df[
                self.summary_df["Сотрудник"] == employee_name
            ].index[0]
            # Add period column if it doesn't exist
            if period not in self.summary_df.columns:
                self.summary_df[period] = 0
            self.summary_df.at[row_idx, period] = total_salary
        else:
            # Create new row
            new_row_data = {"Сотрудник": employee_name}
            # Add zeros for all existing periods
            for existing_period in self.periods:
                new_row_data[existing_period] = (
                    total_salary if existing_period == period else 0
                )

            new_row = pd.DataFrame([new_row_data])
            self.summary_df = pd.concat([self.summary_df, new_row], ignore_index=True)

    def calculate(self, progress_callback=None):
        self.summary_df = pd.DataFrame(columns=["Сотрудник"])
        self.periods = set()

        self.zp_df = self.get_zp_df()
        self.files = self.get_files_df()

        if not self.files:
            print("No files")
            return

        for idx, fl in enumerate(self.files):
            print(f"Processing file: {fl.name}")
            self.calc_zp(fl, self.zp_df, progress_callback, idx)

        # Fill missing period columns with 0 for all employees
        for period in self.periods:
            if period not in self.summary_df.columns:
                self.summary_df[period] = 0

        # Sort by surname (assuming surname is first word in "Сотрудник" column)
        self.summary_df = self.summary_df.sort_values("Сотрудник").reset_index(
            drop=True
        )

        # Save summary file with formulas
        summary_path = self.config.to_files_path / "Ведомость.xlsx"

        # Save using ExcelWriter to add formulas
        with pd.ExcelWriter(summary_path, engine="openpyxl") as writer:
            self.summary_df.to_excel(writer, index=False, sheet_name="Sheet1")

            # Get workbook and worksheet to add sum formulas
            workbook = writer.book
            worksheet = writer.sheets["Sheet1"]

            # Add sum formulas for each period column (skip first column "Сотрудник")
            last_row = len(self.summary_df) + 1  # +1 for header
            for col_idx, col_name in enumerate(
                self.summary_df.columns[1:], start=2
            ):  # Start from column B
                col_letter = worksheet.cell(row=1, column=col_idx).column_letter
                # Add sum formula in the row after last data row
                sum_formula = f"=SUM({col_letter}2:{col_letter}{last_row})"
                worksheet.cell(row=last_row + 1, column=col_idx, value=sum_formula)

            # Add "ИТОГО" label in first column of sum row
            worksheet.cell(row=last_row + 1, column=1, value="ИТОГО")

        # Финальный этап: Сохранение итогового файла
        if progress_callback:
            progress_callback(len(self.files) * 4 + 1)
            
        print(f"Summary saved to: {summary_path}")
